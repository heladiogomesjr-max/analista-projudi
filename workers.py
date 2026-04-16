"""
workers.py — Workers de processamento (XLSX e DJEN)

Cada worker roda em thread separada, atualiza jobs[job_id] e opera
de forma completamente independente.
"""
import os, re, time, shutil, threading
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

import djen
import projudi
import ia
try:
    import sheets as _sheets_mod
    _SHEETS_OK = True
except ImportError:
    _SHEETS_OK = False

# Palavras-chave que indicam falha de site/rede (não de processo)
_ERROS_SITE = ('timeout', 'navigation', 'net::', 'err_connection', 'err_name',
               'target closed', 'target page', 'browser has been closed',
               'page has been closed', 'connection refused', 'socket')

# Erros que indicam que o browser Playwright foi destruído — não adianta aguardar e reconectar
_ERROS_BROWSER_MORTO = ('target page', 'target closed', 'browser has been closed',
                         'page has been closed', 'browser is closed')

PASTA  = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(PASTA, "output")
os.makedirs(OUTPUT, exist_ok=True)

COLUNAS_SAIDA = [
    "NÚMERO DO PROCESSO",
    "DATA DA DECISÃO",
    "DANO MATERIAL",
    "DANO MORAL",
    "TIPO",
    "STATUS DA DECISÃO",
    "RESUMO DO PROCESSO",
    "MATÉRIA",
    "RELATOR/JUIZ",
    "TURMA/VARA",
    "DISTRIBUÍDO 2º GRAU",
    "TEM ACÓRDÃO 2º GRAU",
    "TRANSITADO EM JULGADO? (SIM OU NÃO)",
    "TRANSITADO 1º GRAU",
]

MODELO_XLSX = os.path.join(PASTA, "MODELO.xlsx")


# ══════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ══════════════════════════════════════════════════════════════
_LARGURAS_COL = {
    "NÚMERO DO PROCESSO":                  32,
    "DATA DA DECISÃO":                     16,
    "DANO MATERIAL":                       20,
    "DANO MORAL":                          20,
    "VALOR DA CONDENAÇÃO":                 22,
    "TIPO":                                18,
    "STATUS DA DECISÃO":                   24,
    "RESUMO DO PROCESSO":                  65,
    "MATÉRIA":                             22,
    "RELATOR/JUIZ":                        34,
    "TURMA/VARA":                          28,
    "TRANSITADO EM JULGADO? (SIM OU NÃO)": 12,
    "RESPOSTA IA":                         40,
    "DISTRIBUÍDO 2º GRAU":                 20,
    "TEM ACÓRDÃO 2º GRAU":                 20,
    "TRANSITADO 1º GRAU":                  18,
}
_COLUNAS_WRAP = {"RESUMO DO PROCESSO", "RESPOSTA IA"}


def _nome_arquivo_saida(nome_advogado):
    """Gera nome no formato: RELATÓRIO - NOME SOBRENOME - DD-MM-YYYY HH_MM.xlsx"""
    partes = (nome_advogado or "").strip().split()
    if len(partes) >= 2:
        nome_curto = f"{partes[0]} {partes[1]}"
    elif partes:
        nome_curto = partes[0]
    else:
        nome_curto = "Advogado"
    nome_curto = re.sub(r'[<>:"/\\|?*]', '', nome_curto)
    ts = datetime.now().strftime("%d-%m-%Y %H_%M")
    return f"RELATÓRIO - {nome_curto} - {ts}.xlsx"


def _salvar_xlsx(df, caminho):
    """Salva df usando MODELO.xlsx como template com formatação padronizada.
    Fallback para xlsxwriter caso o modelo não exista."""
    if os.path.exists(MODELO_XLSX):
        shutil.copy2(MODELO_XLSX, caminho)
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        hdrs = [c.value for c in ws[1]]
        for r_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            row_dict = dict(zip(df.columns, row_data))
            ws.row_dimensions[r_idx].height = None  # auto-height
            for c_idx, hdr in enumerate(hdrs, start=1):
                val = row_dict.get(hdr)
                if val is None:
                    for k, v in row_dict.items():
                        if hdr and k and (hdr.startswith(k[:20]) or k.startswith(hdr[:20])):
                            val = v
                            break
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                wrap = hdr in _COLUNAS_WRAP
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=wrap
                )
        for c_idx, hdr in enumerate(hdrs, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = _LARGURAS_COL.get(hdr, 15)

        # Atualiza o range de todas as tabelas do template para cobrir os dados reais
        n_linhas = len(df) + 1  # +1 para o cabeçalho
        n_cols   = len(hdrs)
        novo_ref = f"A1:{get_column_letter(n_cols)}{n_linhas}"
        for tbl in list(ws.tables.values()):
            tbl.ref = novo_ref

        wb.save(caminho)
    else:
        writer = pd.ExcelWriter(caminho, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Processos', index=False)
        wb  = writer.book
        ws  = writer.sheets['Processos']
        max_row, max_col = df.shape
        ws.add_table(0, 0, max_row, max_col - 1, {
            'columns': [{'header': c} for c in df.columns],
            'style': 'Table Style Medium 16',
        })
        fmt_center = wb.add_format({'align': 'center', 'valign': 'vcenter'})
        fmt_wrap   = wb.add_format({'text_wrap': True, 'align': 'center', 'valign': 'top'})
        for i, col in enumerate(df.columns):
            letra = chr(ord('A') + i)
            fmt = fmt_wrap if col in _COLUNAS_WRAP else fmt_center
            ws.set_column(f'{letra}:{letra}', _LARGURAS_COL.get(col, 15), fmt)
        writer.close()


def _montar_linha(numero, dados_projudi, resultado_ia):
    grau         = dados_projudi.get("grau")
    tem_acordao  = grau == 2 and bool(dados_projudi.get("texto_acordao", "").strip())
    return {
        "NÚMERO DO PROCESSO":   numero,
        "DATA DA DECISÃO":      dados_projudi.get("data_decisao", ""),
        "DANO MATERIAL":        resultado_ia.get("DANO_MATERIAL", ""),
        "DANO MORAL":           resultado_ia.get("DANO_MORAL", ""),
        "TIPO":                 dados_projudi.get("tipo", ""),
        "STATUS DA DECISÃO":    resultado_ia.get("DECISAO", ""),
        "RESUMO DO PROCESSO":   resultado_ia.get("RACIOCINIO", ""),
        "MATÉRIA":              resultado_ia.get("MATERIA", ""),
        "RELATOR/JUIZ":         dados_projudi.get("relator_juiz", ""),
        "TURMA/VARA":           dados_projudi.get("turma_vara", ""),
        "DISTRIBUÍDO 2º GRAU":  "SIM" if grau == 2 else "NÃO",
        "TEM ACÓRDÃO 2º GRAU":  "SIM" if tem_acordao else "NÃO",
        "TRANSITADO EM JULGADO? (SIM OU NÃO)":
                                "SIM" if dados_projudi.get("transitado") else "NÃO",
        "TRANSITADO 1º GRAU":   "SIM" if dados_projudi.get("transitado_1g") else "NÃO",
    }


def _detectar_relator_designado(texto_acordao):
    """
    Detecta o padrão de relator designado (voto vencedor) no texto do acórdão.
    Retorna (relator_vencido, relator_designado) ou ("", "") se não encontrado.

    Padrão esperado no PROJUDI:
      "... Fulano de Tal (relator vencido) e Ciclano da Silva (relator designado)."
    """
    if not texto_acordao:
        return "", ""

    # Palavras que aparecem antes do nome do juiz mas não fazem parte dele
    _TITULOS = {'Juiz', 'Juízes', 'Juíza', 'Juízas', 'Des', 'Dr', 'Dra',
                'Desembargador', 'Desembargadora', 'Magistrado', 'Magistrada'}

    def _nome_antes(texto, marcador):
        """Extrai o nome próprio imediatamente antes do marcador."""
        idx = texto.lower().find(marcador.lower())
        if idx == -1:
            return ""
        trecho = texto[max(0, idx - 150):idx]
        palavras = trecho.split()
        nome = []
        for p in reversed(palavras):
            p_limpa = re.sub(r'[^\wÀ-ú]', '', p)
            if not p_limpa:
                continue
            if re.match(r'^[A-ZÀ-Ú]', p_limpa):        # palavra capitalizada
                nome.insert(0, p_limpa)
            elif re.match(r'^d[aeo]s?$', p_limpa, re.I): # conector: da/do/de/dos/das
                nome.insert(0, p_limpa)
            else:
                break                                     # lowercase comum → para
        # Remove títulos que precedem o nome mas não fazem parte dele
        while nome and nome[0] in _TITULOS:
            nome.pop(0)
        return ' '.join(nome)

    vencido   = _nome_antes(texto_acordao, '(relator vencido)')   \
             or _nome_antes(texto_acordao, '(relatora vencida)')
    designado = _nome_antes(texto_acordao, '(relator designado)') \
             or _nome_antes(texto_acordao, '(relatora designada)')
    return vencido, designado


class _CanceladoError(Exception):
    """Levantada quando o usuário cancela durante o processamento."""


def _is_erro_site(e):
    """Retorna True se o erro é de indisponibilidade do site (não de processo)."""
    return isinstance(e, PWTimeout) or any(k in str(e).lower() for k in _ERROS_SITE)


def _reconectar(page, cpf, senha, log, espera=60, max_tentativas=10, job=None):
    """Aguarda o PROJUDI voltar, faz login e retorna (url_2g, url_1g) atualizadas."""
    for t in range(1, max_tentativas + 1):
        log(f"   ⏳ Site indisponível — aguardando {espera}s... (tentativa {t}/{max_tentativas})")
        # Sleep interrompível: verifica cancelamento a cada 0.5s
        for _ in range(espera * 2):
            if job and job.get('cancelado'):
                raise _CanceladoError()
            time.sleep(0.5)
        if job and job.get('cancelado'):
            raise _CanceladoError()
        try:
            projudi.login(page, cpf, senha, log)
            url_2g, url_1g = projudi.get_urls_busca(page, log)
            log("   ✅ Reconectado ao PROJUDI. Retomando processamento...")
            return url_2g, url_1g
        except _CanceladoError:
            raise
        except Exception as e_recon:
            # Browser morto (OOM, kill externo) — não adianta aguardar e tentar de novo
            msg = str(e_recon).lower()
            if any(k in msg for k in _ERROS_BROWSER_MORTO):
                raise RuntimeError(
                    f"Browser Playwright encerrado inesperadamente (possível falta de memória). "
                    f"Detalhe: {e_recon}"
                )
            log(f"   ⚠️ Falha ao reconectar ({e_recon}). Tentando novamente...")
    raise RuntimeError(f"PROJUDI permanece indisponível após {max_tentativas} tentativas de reconexão.")


def _processar_com_retry(page, numero, url_2g_box, url_1g_box,
                         cpf, senha, api_key, log, modelo_ia=None,
                         nome_advogado=None, max_tentativas=10, espera=60,
                         relator_filtro=None, usar_ia=True, job=None):
    """
    Tenta processar o número com retry automático quando o site cai.
    url_2g_box / url_1g_box são listas de 1 elemento (mutáveis) para permitir
    atualização das URLs após reconexão.
    """
    for tentativa in range(1, max_tentativas + 1):
        try:
            return _processar_numero(page, numero, url_2g_box[0], url_1g_box[0],
                                     api_key, log, modelo_ia, nome_advogado,
                                     relator_filtro, usar_ia=usar_ia)
        except _CanceladoError:
            raise
        except Exception as e:
            # Se o browser foi fechado pelo monitor de cancelamento, converte em _CanceladoError
            if job and job.get('cancelado'):
                raise _CanceladoError()
            if _is_erro_site(e) and tentativa < max_tentativas:
                log(f"   ❌ Erro de site: {e}")
                url_2g_box[0], url_1g_box[0] = _reconectar(
                    page, cpf, senha, log, espera=espera,
                    max_tentativas=max_tentativas, job=job,
                )
                log(f"   🔄 Retentando processo {numero}...")
            else:
                raise


def _processar_numero(page, numero, url_2g, url_1g, api_key, log, modelo_ia=None,
                      nome_advogado=None, relator_filtro=None, usar_ia=True):
    """Pipeline completo para um processo: extrai no PROJUDI + classifica com IA."""
    dados = projudi.analisar_processo(
        page, numero, url_2g, url_1g, log,
        extrair_textos=usar_ia, relator_filtro=relator_filtro,
    )

    if dados.get("ignorado"):
        return {"_ignorado": True, "NÚMERO DO PROCESSO": numero}

    if dados["tipo"] == "NÃO LOCALIZADO":
        return _montar_linha(numero, dados, {})

    if not usar_ia:
        log(f"   IA desativada.")
        return _montar_linha(numero, dados, {})

    sem_acordao  = not dados.get("texto_acordao", "").strip()
    sem_sentenca = not dados.get("texto_sentenca", "").strip()
    if sem_acordao and sem_sentenca:
        log(f"   ⏭️ Sem acórdão nem sentença — processo pulado.")
        return _montar_linha(numero, dados, {})

    tipo                    = dados["tipo"]
    turma_vara              = dados["turma_vara"]
    relator_juiz            = dados["relator_juiz"]
    texto_acordao           = dados["texto_acordao"]
    texto_acordao_embargos  = dados.get("texto_acordao_embargos", "")
    texto_sentenca          = dados["texto_sentenca"]
    texto_sentenca_embargos = dados.get("texto_sentenca_embargos", "")
    texto_peticao           = dados["texto_peticao"]

    # Detecta relator designado (voto vencedor): corrige o relator antes de enviar à IA
    vencido, designado = _detectar_relator_designado(texto_acordao)
    if designado:
        log(f"   ⚖️ Relator vencido: {vencido or relator_juiz} → Relator designado: {designado}")
        relator_juiz        = designado
        dados["relator_juiz"] = designado  # garante que _montar_linha também usa o correto

    if tipo == "ACÓRDÃO" and texto_acordao.strip():
        texto_principal          = texto_acordao
        texto_embargos_principal = texto_acordao_embargos
    elif tipo == "ACÓRDÃO" and not texto_acordao.strip():
        # Distribuído no 2º grau mas sem acórdão publicado ainda — analisa sentença do 1º grau
        tipo                     = "SENTENÇA"
        texto_principal          = texto_sentenca
        texto_embargos_principal = texto_sentenca_embargos
        log("   2g sem acórdão — analisando sentença do 1º grau.")
        # Usa o juiz e vara do 1º grau (não o relator do 2g)
        juiz_1g = dados.get("juiz_sentenca", "")
        vara_1g = dados.get("vara_sentenca", "")
        if juiz_1g:
            relator_juiz          = juiz_1g
            turma_vara            = vara_1g
            dados["relator_juiz"] = juiz_1g
            dados["turma_vara"]   = vara_1g
        dados["tipo"] = "SENTENÇA"  # garante que _montar_linha registra SENTENÇA
    else:
        texto_principal          = texto_sentenca
        texto_embargos_principal = ""

    partes = ia.extrair_partes(texto_principal or texto_sentenca or texto_peticao)

    resultado_ia = ia.classificar(
        numero, tipo, turma_vara, relator_juiz,
        partes, texto_principal, texto_sentenca, texto_peticao,
        api_key, log, model=modelo_ia, nome_advogado=nome_advogado or "",
        texto_embargos_principal=texto_embargos_principal,
        texto_sentenca_embargos=texto_sentenca_embargos,
    )
    log(f"   ✅ {tipo} | {resultado_ia.get('DECISAO','')} | {resultado_ia.get('MATERIA','')}")

    return _montar_linha(numero, dados, resultado_ia)


# ══════════════════════════════════════════════════════════════
# PIPELINE UNIFICADO
# ══════════════════════════════════════════════════════════════
def _executar_pipeline(job_id, jobs, numeros, cpf, senha, api_key,
                       modelo_ia, nome_advogado, usar_ia, relator_filtro=None):
    """
    Pipeline unificado: processa lista de números no PROJUDI.
    Suporta pause/cancel e salva checkpoint após cada processo.
    """
    job   = jobs[job_id]
    total = len(numeros)
    nome_out = os.path.join(OUTPUT, _nome_arquivo_saida(nome_advogado))
    job['file'] = nome_out  # disponibiliza desde o início para checkpoint

    def log(msg):
        print(msg, flush=True)
        job['logs'].append(msg)

    def pct(p, sub=''):
        job['pct'] = p
        job['subtitulo'] = sub

    pct(8, "Abrindo navegador...")
    linhas = []
    _pipeline_ativa = threading.Event()
    _pipeline_ativa.set()

    try:
        with sync_playwright() as pw:
            browser, page = projudi.novo_browser(pw)

            # ── Thread monitor: fecha o browser imediatamente ao cancelar ──
            def _monitor_cancel():
                while _pipeline_ativa.is_set():
                    if job.get('cancelado'):
                        try:
                            browser.close()
                        except Exception:
                            pass
                        break
                    time.sleep(0.2)

            threading.Thread(target=_monitor_cancel, daemon=True).start()

            try:
                pct(10, "Fazendo login...")
                projudi.login(page, cpf, senha, log)

                pct(14, "Localizando menus...")
                url_2g, url_1g = projudi.get_urls_busca(page, log)
                url_2g_box, url_1g_box = [url_2g], [url_1g]

                for idx, numero in enumerate(numeros, 1):
                    # Verifica cancelamento
                    if job.get('cancelado'):
                        log("⛔ Processamento cancelado.")
                        break

                    # Verifica pausa
                    if job.get('pausado'):
                        log("⏸️ Processamento pausado.")
                        while job.get('pausado') and not job.get('cancelado'):
                            time.sleep(0.5)
                        if job.get('cancelado'):
                            log("⛔ Processamento cancelado.")
                            break
                        log("▶️ Processamento retomado.")

                    pct(14 + int((idx / total) * 80), f"Processo {idx}/{total}")
                    log(f"\n[{idx}/{total}] {numero}")
                    try:
                        linha = _processar_com_retry(
                            page, numero, url_2g_box, url_1g_box,
                            cpf, senha, api_key, log, modelo_ia,
                            nome_advogado, usar_ia=usar_ia,
                            relator_filtro=relator_filtro,
                            job=job,
                        )
                    except _CanceladoError:
                        log("⛔ Processamento cancelado.")
                        break
                    except Exception as e:
                        log(f"   ❌ Erro: {e}")
                        linha = {c: "" for c in COLUNAS_SAIDA}
                        linha.update({"NÚMERO DO PROCESSO": numero, "TIPO": "ERRO",
                                      "RESUMO DO PROCESSO": str(e)})

                    if linha.get("_ignorado"):
                        continue
                    linhas.append(linha)
                    job['linhas'] = linhas[:]  # snapshot para o dashboard

                    # Checkpoint: salva planilha após cada processo concluído
                    try:
                        df_cp = pd.DataFrame(linhas, columns=COLUNAS_SAIDA)
                        _salvar_xlsx(df_cp, nome_out)
                    except Exception as _e_cp:
                        log(f"   ⚠️ Erro ao salvar checkpoint: {_e_cp}")

                    # Envia para Google Sheets (se configurado)
                    if _SHEETS_OK and not linha.get("TIPO") in ("NÃO LOCALIZADO", "ERRO"):
                        try:
                            _sheets_mod.inserir_na_planilha(
                                [linha],
                                turma_vara=linha.get("TURMA/VARA") or "Geral",
                                advogado_key=job.get("advogado_key"),
                                log=log,
                            )
                        except Exception as _e_sh:
                            log(f"   ⚠️ Sheets: {_e_sh}")

                    time.sleep(1.5)
            finally:
                _pipeline_ativa.clear()
                try:
                    browser.close()
                except Exception:
                    pass

    except AttributeError as e:
        # Playwright lança AttributeError no teardown quando o browser é fechado
        # externamente pelo monitor de cancelamento — comportamento esperado.
        if '_playwright' not in str(e):
            raise

    return linhas


def _finalizar_job(job_id, jobs, linhas, api_key, modelo_ia, usar_ia, nome_advogado=""):
    """Gera relatório analítico e marca o job como concluído ou cancelado."""
    job = jobs[job_id]

    def log(msg):
        print(msg, flush=True)
        job['logs'].append(msg)

    def pct(p, sub=''):
        job['pct'] = p
        job['subtitulo'] = sub

    ok = sum(1 for l in linhas if l.get("STATUS DA DECISÃO"))
    log(f"\n✅ {ok}/{len(linhas)} processos classificados.")

    if usar_ia and api_key and ok >= 2:
        pct(97, "Gerando relatório analítico...")
        texto_rel = ia.gerar_relatorio(linhas, api_key, modelo_ia or ia.MODELO_PADRAO, log)
        if texto_rel:
            # Mesmo padrão de nome do XLSX: RELATÓRIO - NOME - DD-MM-YYYY HH_MM.docx
            base_xlsx = os.path.basename(job.get('file', ''))
            nome_docx_base = base_xlsx.replace('.xlsx', '.docx') if base_xlsx else f"relatorio_{job_id}.docx"
            nome_docx = os.path.join(OUTPUT, nome_docx_base)
            if ia.gerar_docx(texto_rel, nome_docx, total_processos=ok, nome_advogado=nome_advogado):
                job['docx_file'] = nome_docx
                log("   ✅ Relatório analítico gerado.")
            else:
                log("   ⚠️ python-docx não instalado. Execute: pip install python-docx")

    job['status'] = 'cancelled' if job.get('cancelado') else 'done'
    pct(100, "Pronto!")


# ══════════════════════════════════════════════════════════════
# WORKER XLSX
# ══════════════════════════════════════════════════════════════
def processar_job_xlsx(job_id, jobs, caminho_xlsx, cpf, senha, api_key, batch_size,
                       modelo_ia=None, nome_advogado=None, usar_ia=True,
                       numeros_texto="", relator_filtro=None, advogado_key=None):
    job = jobs[job_id]
    job['advogado_key'] = advogado_key or 'luis_albert'

    def log(msg):
        print(msg, flush=True)
        job['logs'].append(msg)

    def pct(p, sub=''):
        job['pct'] = p
        job['subtitulo'] = sub

    try:
        ia.limpar_cache()

        # 1. Obtém lista de números — textarea tem prioridade sobre arquivo
        if numeros_texto and numeros_texto.strip():
            numeros = [p.strip() for p in numeros_texto.splitlines() if p.strip()]
        elif caminho_xlsx and os.path.exists(caminho_xlsx):
            df = pd.read_excel(caminho_xlsx, dtype=str).fillna("")
            col_proc = next((c for c in df.columns if "processo" in c.lower()), None)
            if not col_proc:
                job['status'] = 'error'
                job['error']  = "Coluna de processo não encontrada. Certifique-se de que existe uma coluna com 'PROCESSO' no nome."
                return
            numeros = [p.strip() for p in df[col_proc].tolist() if p.strip()]
        else:
            job['status'] = 'error'
            job['error']  = "Nenhum arquivo enviado e nenhum processo informado na caixa de texto."
            return

        antes = len(numeros)
        numeros = list(dict.fromkeys(numeros))
        if len(numeros) < antes:
            log(f"⚠️ {antes - len(numeros)} processo(s) duplicado(s) removido(s) ({antes} → {len(numeros)}).")
        if batch_size:
            numeros = numeros[:batch_size]
        log(f"✅ {len(numeros)} processos para analisar.")
        pct(6, "Lista de processos pronta.")

        # 2. Pipeline unificado
        linhas = _executar_pipeline(
            job_id, jobs, numeros, cpf, senha, api_key,
            modelo_ia, nome_advogado, usar_ia, relator_filtro,
        )

        # 3. Finaliza (relatório + status)
        _finalizar_job(job_id, jobs, linhas, api_key, modelo_ia, usar_ia, nome_advogado=nome_advogado or "")

    except Exception as e:
        job['status'] = 'error'
        job['error']  = str(e)
        log(f"❌ Erro geral: {e}")


# ══════════════════════════════════════════════════════════════
# WORKER DJEN
# ══════════════════════════════════════════════════════════════
def processar_job_djen(job_id, jobs, nome_adv, data_ini, data_fim, turma,
                       relator_filtro, cpf, senha, api_key, batch_size,
                       filtro_texto='', modelo_ia=None, nome_advogado=None, usar_ia=True,
                       advogado_key=None, filtro_tipo_doc=False):
    job = jobs[job_id]
    job['advogado_key'] = advogado_key or 'luis_albert'
    _t0 = time.time()

    def log(msg):
        elapsed = time.time() - _t0
        msg_t = f"[{elapsed:6.1f}s] {msg}"
        print(msg_t, flush=True)
        job['logs'].append(msg_t)

    def pct(p, sub=''):
        job['pct'] = p
        job['subtitulo'] = sub

    try:
        ia.limpar_cache()

        # 1. Busca no DJEN
        pct(5, "Buscando no DJEN...")
        filtro_rel = f" | Relator: {relator_filtro}" if relator_filtro else ""
        log(f"DJEN: {nome_adv} | {data_ini} → {data_fim} | Turma: {turma or 'Todas'}{filtro_rel}")

        # Tipos de acórdão confirmados na API do TJAM:
        #   COM JULGAMENTO DE MÉRITO / SEM JULGAMENTO DE MÉRITO — mais comuns no TJAM
        #   JUNTADA DE ACÓRDÃO / ACÓRDÃO ... — nomenclatura de outros tribunais PJe
        _TIPOS_ACORDAO = {
            'JUNTADA', 'COM JULGAMENTO', 'SEM JULGAMENTO',
            'ACÓRDÃO', 'ACORDAO',
        }

        if filtro_tipo_doc:
            # Publicações de acórdão são indexadas sob orgaoId diferente dos órgãos selecionados.
            # Buscar sem orgaoId (global) e filtrar turma + tipo client-side evita o problema de
            # deduplicação onde distribuições de processo bloqueiam os acórdãos do mesmo processo.
            log(f"   ℹ️  Modo acórdão: busca global + filtro de turma e tipo client-side")
            processos_djen = djen.buscar(nome_adv, data_ini, data_fim, '0', log=log)

            # Filtra pela turma selecionada (se não for 'todas')
            if turma and turma != '0':
                orgao_ids_sel = djen._resolver_orgaos(turma)
                nomes_turmas  = {djen._NOMES_ORGAOS.get(oid, '').upper()
                                 for oid in orgao_ids_sel} - {''}
                if nomes_turmas:
                    antes = len(processos_djen)
                    processos_djen = [p for p in processos_djen
                                      if p.get('turma_djen', '') in nomes_turmas]
                    log(f"   🔍 Filtro turma: {len(processos_djen)}/{antes} publicações mantidas.")
        else:
            processos_djen = djen.buscar(nome_adv, data_ini, data_fim, turma or '0', log=log)

        if not processos_djen:
            job['status'] = 'error'
            job['error']  = "Nenhum processo encontrado no DJEN com esses parâmetros."
            return

        # Filtro por palavra-chave no texto da publicação
        if filtro_texto:
            ft = filtro_texto.lower()
            antes = len(processos_djen)
            processos_djen = [p for p in processos_djen if ft in p.get('texto', '').lower()]
            log(f"🔍 Filtro '{filtro_texto}': {len(processos_djen)}/{antes} publicações mantidas.")
            if not processos_djen:
                job['status'] = 'error'
                job['error']  = f"Nenhuma publicação contém a palavra-chave '{filtro_texto}'."
                return

        if filtro_tipo_doc:
            antes = len(processos_djen)
            processos_djen = [p for p in processos_djen
                              if any(kw in p.get('tipo_doc', '').upper()
                                     for kw in _TIPOS_ACORDAO)]
            log(f"🔍 Filtro acórdão: {len(processos_djen)}/{antes} publicações mantidas.")
            if not processos_djen:
                job['status'] = 'error'
                job['error']  = "Nenhuma publicação de acórdão encontrada no período."
                return

        numeros = [p['PROCESSO'] for p in processos_djen]
        antes = len(numeros)
        numeros = list(dict.fromkeys(numeros))
        if len(numeros) < antes:
            log(f"⚠️ {antes - len(numeros)} processo(s) duplicado(s) removido(s) ({antes} → {len(numeros)}).")
        if batch_size:
            numeros = numeros[:batch_size]
        log(f"✅ {len(numeros)} processos para analisar.")
        pct(6, "Lista de processos pronta.")

        # 2. Pipeline unificado
        linhas = _executar_pipeline(
            job_id, jobs, numeros, cpf, senha, api_key,
            modelo_ia, nome_advogado, usar_ia, relator_filtro,
        )

        # 3. Finaliza (relatório + status)
        _finalizar_job(job_id, jobs, linhas, api_key, modelo_ia, usar_ia, nome_advogado=nome_advogado or "")

    except Exception as e:
        job['status'] = 'error'
        job['error']  = str(e)
        log(f"❌ Erro geral: {e}")
