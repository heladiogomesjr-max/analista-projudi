"""
teste_prompt.py — Script de validação do prompt de IA.

Processa uma lista fixa de números no PROJUDI, classifica com IA e salva
XLSX de validação com colunas extras para revisão manual.

Uso:
  python teste_prompt.py                  → usa CPF/senha do [projudi] no config.ini
  python teste_prompt.py LUIS             → usa credenciais do usuário LUIS
  python teste_prompt.py --limpar-cache   → limpa cache IA antes de rodar
  python teste_prompt.py LUIS --limpar-cache
"""
import sys, re, os, time, configparser
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# Força UTF-8 no console do Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent))
import projudi
import ia

# ══════════════════════════════════════════════════════════════
# LISTA DE PROCESSOS PARA VALIDAÇÃO
# ══════════════════════════════════════════════════════════════
PROCESSOS = [
    "0064550-86.2024.8.04.1000",
    "0105052-67.2024.8.04.1000",
    "0601226-24.2025.8.04.4400",
    "0221520-80.2025.8.04.1000",
    "0284295-34.2025.8.04.1000",
    "0000943-97.2025.8.04.2800",
    "0000876-35.2025.8.04.2800",
    "0655423-41.2025.8.04.1000",
    "0685708-17.2025.8.04.1000",
    "0601219-32.2025.8.04.4400",
    "0601328-46.2025.8.04.4400",
]

INI    = Path(__file__).parent / "config.ini"
OUTPUT = Path(__file__).parent / "output"
OUTPUT.mkdir(exist_ok=True)

# Erros de rede/site (para reconexão)
_ERROS_SITE = ('timeout', 'navigation', 'net::', 'err_connection',
               'target closed', 'target page', 'browser has been closed',
               'page has been closed', 'socket')


# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════
def _ler_config(filtro_usuario=""):
    cfg = configparser.ConfigParser()
    cfg.read(INI, encoding="utf-8")

    api_key    = cfg.get("claude",   "api_key", fallback="")
    modelo_ia  = cfg.get("claude",   "modelo",  fallback="claude-haiku-4-5-20251001")
    nome_adv   = cfg.get("advogado", "nome",    fallback="")

    if not filtro_usuario:
        cpf   = cfg.get("projudi", "cpf",   fallback="")
        senha = cfg.get("projudi", "senha", fallback="")
        label = "projudi"
    else:
        filtro_up = filtro_usuario.upper()
        cpf = senha = label = ""
        i = 0
        while cfg.has_option("usuarios", f"cpf_{i}"):
            lbl = cfg.get("usuarios", f"label_{i}", fallback="")
            if filtro_up in lbl.upper():
                cpf   = cfg.get("usuarios", f"cpf_{i}",   fallback="")
                senha = cfg.get("usuarios", f"senha_{i}", fallback="")
                label = lbl
                break
            i += 1
        if not cpf:
            print(f"❌ Usuário '{filtro_usuario}' não encontrado no config.ini.")
            sys.exit(1)

    return cpf, senha, label, api_key, modelo_ia, nome_adv


# ══════════════════════════════════════════════════════════════
# RELATOR DESIGNADO (replicado de workers.py)
# ══════════════════════════════════════════════════════════════
_TITULOS = {'Juiz', 'Juízes', 'Juíza', 'Juízas', 'Des', 'Dr', 'Dra',
            'Desembargador', 'Desembargadora', 'Magistrado', 'Magistrada'}

def _detectar_relator_designado(texto_acordao):
    if not texto_acordao:
        return "", ""

    def _nome_antes(texto, marcador):
        idx = texto.lower().find(marcador.lower())
        if idx == -1:
            return ""
        trecho = texto[max(0, idx - 150):idx]
        palavras = trecho.split()
        nome = []
        for p in reversed(palavras):
            p_limpa = re.sub(r'[^\wÀ-ú]', '', p)
            if not p_limpa:
                continue
            if re.match(r'^[A-ZÀ-Ú]', p_limpa):
                nome.insert(0, p_limpa)
            elif re.match(r'^d[aeo]s?$', p_limpa, re.I):
                nome.insert(0, p_limpa)
            else:
                break
        while nome and nome[0] in _TITULOS:
            nome.pop(0)
        return ' '.join(nome)

    vencido   = (_nome_antes(texto_acordao, '(relator vencido)')
              or _nome_antes(texto_acordao, '(relatora vencida)'))
    designado = (_nome_antes(texto_acordao, '(relator designado)')
              or _nome_antes(texto_acordao, '(relatora designada)'))
    return vencido, designado


# ══════════════════════════════════════════════════════════════
# PIPELINE DE UM PROCESSO
# ══════════════════════════════════════════════════════════════
def _processar(page, numero, url_2g, url_1g, api_key, modelo_ia, nome_adv, log):
    dados = projudi.analisar_processo(
        page, numero, url_2g, url_1g, log, extrair_textos=True,
    )

    if dados["tipo"] == "NÃO LOCALIZADO":
        log("   ⚠️ Não localizado.")
        return _montar_linha(numero, dados, {})

    tipo                    = dados["tipo"]
    turma_vara              = dados["turma_vara"]
    relator_juiz            = dados["relator_juiz"]
    texto_acordao           = dados.get("texto_acordao", "")
    texto_acordao_embargos  = dados.get("texto_acordao_embargos", "")
    texto_sentenca          = dados.get("texto_sentenca", "")
    texto_sentenca_embargos = dados.get("texto_sentenca_embargos", "")
    texto_peticao           = dados.get("texto_peticao", "")

    # Detecta relator designado (voto por maioria)
    vencido, designado = _detectar_relator_designado(texto_acordao)
    if designado:
        log(f"   ⚖️ Relator vencido: {vencido or relator_juiz} → Designado: {designado}")
        relator_juiz          = designado
        dados["relator_juiz"] = designado

    # Define documento principal
    if tipo == "ACÓRDÃO" and texto_acordao.strip():
        texto_principal          = texto_acordao
        texto_embargos_principal = texto_acordao_embargos
    elif tipo == "ACÓRDÃO" and not texto_acordao.strip():
        tipo                     = "SENTENÇA"
        texto_principal          = texto_sentenca
        texto_embargos_principal = texto_sentenca_embargos
        dados["tipo"]            = "SENTENÇA"
        juiz_1g = dados.get("juiz_sentenca", "")
        vara_1g = dados.get("vara_sentenca", "")
        if juiz_1g:
            relator_juiz          = juiz_1g
            turma_vara            = vara_1g
            dados["relator_juiz"] = juiz_1g
            dados["turma_vara"]   = vara_1g
        log("   2g sem acórdão — analisando sentença do 1º grau.")
    else:
        texto_principal          = texto_sentenca
        texto_embargos_principal = ""

    if not texto_principal.strip() and not texto_sentenca.strip():
        log("   ⏭️ Sem texto extraído — pulando IA.")
        return _montar_linha(numero, dados, {})

    partes = ia.extrair_partes(texto_principal or texto_sentenca or texto_peticao)

    resultado_ia = ia.classificar(
        numero, tipo, turma_vara, relator_juiz,
        partes, texto_principal, texto_sentenca, texto_peticao,
        api_key, log,
        model=modelo_ia,
        nome_advogado=nome_adv,
        texto_embargos_principal=texto_embargos_principal,
        texto_sentenca_embargos=texto_sentenca_embargos,
        texto_movimentos=dados.get("texto_movimentos", ""),
    )

    decisao  = resultado_ia.get("DECISAO", "")
    materia  = resultado_ia.get("MATERIA", "")
    log(f"   ✅ {tipo} | {decisao} | {materia}")

    return _montar_linha(numero, dados, resultado_ia)


def _montar_linha(numero, dados, resultado_ia):
    grau        = dados.get("grau")
    tem_acordao = grau == 2 and bool(dados.get("texto_acordao", "").strip())
    return {
        "NÚMERO DO PROCESSO":                   numero,
        "DATA DA DECISÃO":                      dados.get("data_decisao", ""),
        "TIPO":                                 dados.get("tipo", ""),
        "STATUS DA DECISÃO":                    resultado_ia.get("DECISAO", ""),
        "MATÉRIA":                              resultado_ia.get("MATERIA", ""),
        "DANO MORAL":                           resultado_ia.get("DANO_MORAL", ""),
        "DANO MATERIAL":                        resultado_ia.get("DANO_MATERIAL", ""),
        "TRANSITADO EM JULGADO? (SIM OU NÃO)":
            resultado_ia.get("TRANSITADO") or ("SIM" if dados.get("transitado") else "NÃO"),
        "RELATOR/JUIZ":                         dados.get("relator_juiz", ""),
        "TURMA/VARA":                           dados.get("turma_vara", ""),
        "DISTRIBUÍDO 2º GRAU":                  "SIM" if grau == 2 else "NÃO",
        "TEM ACÓRDÃO 2º GRAU":                  "SIM" if tem_acordao else "NÃO",
        # Colunas de validação — preenchidas manualmente pelo revisor
        "RACIOCÍNIO IA":                        resultado_ia.get("RACIOCINIO", ""),
        "VALIDAÇÃO":                            "",   # CORRETO / ERRADO
        "ERRO — STATUS":                        "",   # qual deveria ser o status correto
        "ERRO — MATÉRIA":                       "",   # qual deveria ser a matéria correta
        "OBSERVAÇÃO":                           "",   # descrição livre do erro
    }


# ══════════════════════════════════════════════════════════════
# SALVAR XLSX DE VALIDAÇÃO
# ══════════════════════════════════════════════════════════════
def _salvar_xlsx_validacao(linhas, caminho):
    try:
        import openpyxl
        from openpyxl.styles import (Alignment, PatternFill, Font,
                                      Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("   ⚠️ openpyxl não instalado — pulando XLSX.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validação Prompt"

    if not linhas:
        wb.save(caminho)
        return

    colunas = list(linhas[0].keys())

    # Cores para cabeçalho
    _COR_HEADER_DADOS  = "1F497D"   # azul escuro — colunas de dados
    _COR_HEADER_VALID  = "833C00"   # marrom escuro — colunas de validação
    _COR_STATUS = {
        "FAVORÁVEL":            "C6EFCE",
        "DESFAVORÁVEL":         "FFC7CE",
        "SENTENÇA ANULADA":     "FFEB9C",
        "EXTINTO SEM MÉRITO":   "D9D9D9",
        "ACORDO HOMOLOGADO":    "BDD7EE",
        "SEM PARECER CONCLUSIVO": "E2EFDA",
    }

    LARGURAS = {
        "NÚMERO DO PROCESSO":                  28,
        "DATA DA DECISÃO":                     14,
        "TIPO":                                14,
        "STATUS DA DECISÃO":                   22,
        "MATÉRIA":                             22,
        "DANO MORAL":                          14,
        "DANO MATERIAL":                       14,
        "TRANSITADO EM JULGADO? (SIM OU NÃO)": 12,
        "RELATOR/JUIZ":                        30,
        "TURMA/VARA":                          26,
        "DISTRIBUÍDO 2º GRAU":                 16,
        "TEM ACÓRDÃO 2º GRAU":                 16,
        "RACIOCÍNIO IA":                       70,
        "VALIDAÇÃO":                           16,
        "ERRO — STATUS":                       22,
        "ERRO — MATÉRIA":                      22,
        "OBSERVAÇÃO":                          40,
    }

    COLUNAS_VALIDACAO = {"VALIDAÇÃO", "ERRO — STATUS", "ERRO — MATÉRIA", "OBSERVAÇÃO"}

    # Cabeçalho
    for c_idx, col in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        is_valid = col in COLUNAS_VALIDACAO
        cor = _COR_HEADER_VALID if is_valid else _COR_HEADER_DADOS
        cell.fill   = PatternFill("solid", fgColor=cor)
        cell.font   = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.column_dimensions[get_column_letter(c_idx)].width = LARGURAS.get(col, 15)

    ws.row_dimensions[1].height = 32

    thin = Side(style="thin", color="CCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Dados
    for r_idx, linha in enumerate(linhas, 2):
        status = linha.get("STATUS DA DECISÃO", "")
        cor_status = _COR_STATUS.get(status, "FFFFFF")

        for c_idx, col in enumerate(colunas, 1):
            val  = linha.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = borda

            wrap = col in ("RACIOCÍNIO IA", "OBSERVAÇÃO")
            cell.alignment = Alignment(horizontal="left" if wrap else "center",
                                       vertical="top" if wrap else "center",
                                       wrap_text=wrap)

            # Colorir linha conforme status da decisão (apenas colunas de dados)
            if col not in COLUNAS_VALIDACAO and col != "RACIOCÍNIO IA":
                cell.fill = PatternFill("solid", fgColor=cor_status)

        ws.row_dimensions[r_idx].height = None  # auto

    # Congela cabeçalho
    ws.freeze_panes = "A2"

    wb.save(caminho)
    print(f"   💾 XLSX salvo: {caminho}")


# ══════════════════════════════════════════════════════════════
# RESUMO NO CONSOLE
# ══════════════════════════════════════════════════════════════
def _imprimir_resumo(linhas):
    print("\n" + "═" * 90)
    print(f"  {'PROCESSO':<32} {'TIPO':<12} {'STATUS':<26} {'MATÉRIA':<22} {'TRANS.'}")
    print("─" * 90)
    for l in linhas:
        num     = l.get("NÚMERO DO PROCESSO", "")
        tipo    = l.get("TIPO", "")[:10]
        status  = l.get("STATUS DA DECISÃO", "")[:24]
        materia = l.get("MATÉRIA", "")[:20]
        trans   = l.get("TRANSITADO EM JULGADO? (SIM OU NÃO)", "")
        print(f"  {num:<32} {tipo:<12} {status:<26} {materia:<22} {trans}")
    print("═" * 90)

    # Contagem de status
    from collections import Counter
    cnt = Counter(l.get("STATUS DA DECISÃO", "N/A") for l in linhas)
    print("\n  RESUMO:")
    for status, qtd in sorted(cnt.items()):
        print(f"    {status:<30} {qtd} processo(s)")
    print()


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    args           = sys.argv[1:]
    limpar_cache   = "--limpar-cache" in args
    filtro_usuario = next((a for a in args if not a.startswith("--")), "")

    cpf, senha, label, api_key, modelo_ia, nome_adv = _ler_config(filtro_usuario)

    if not cpf or not senha:
        print("❌ CPF ou senha vazio no config.ini.")
        sys.exit(1)
    if not api_key:
        print("❌ api_key Claude vazia no config.ini.")
        sys.exit(1)

    if limpar_cache:
        ia.limpar_cache()
        print("🗑️  Cache IA limpo.\n")

    cpf_mask = cpf[:3] + ".***.***-" + cpf[-2:] if len(re.sub(r'\D','',cpf)) == 11 else cpf
    ts_inicio = datetime.now().strftime("%d/%m/%Y %H:%M")

    print(f"\n{'═'*60}")
    print(f"  TESTE DE VALIDAÇÃO DO PROMPT — {ts_inicio}")
    print(f"  Usuário  : {label}  ({cpf_mask})")
    print(f"  Modelo IA: {modelo_ia}")
    print(f"  Processos: {len(PROCESSOS)}")
    print(f"{'═'*60}\n")

    linhas = []

    with sync_playwright() as pw:
        browser, page = projudi.novo_browser(pw)

        try:
            def log(msg):
                print(msg, flush=True)

            projudi.login(page, cpf, senha, log)
            url_2g, url_1g = projudi.get_urls_busca(page, log)

            if not url_2g and not url_1g:
                print("❌ URLs de busca não encontradas — verifique o login.")
                sys.exit(1)

            print(f"\n✅ Login OK. Iniciando {len(PROCESSOS)} processos...\n")

            for idx, numero in enumerate(PROCESSOS, 1):
                print(f"\n{'─'*60}")
                print(f"  [{idx}/{len(PROCESSOS)}] {numero}")
                print(f"{'─'*60}")

                for tentativa in range(1, 4):
                    try:
                        linha = _processar(
                            page, numero, url_2g, url_1g,
                            api_key, modelo_ia, nome_adv, log,
                        )
                        linhas.append(linha)
                        break
                    except PWTimeout as e:
                        if tentativa < 3:
                            print(f"   ⏳ Timeout — reconectando ({tentativa}/3)...")
                            try:
                                projudi.login(page, cpf, senha, log)
                                url_2g, url_1g = projudi.get_urls_busca(page, log)
                            except Exception:
                                pass
                        else:
                            print(f"   ❌ Falha após 3 tentativas: {e}")
                            linhas.append(_montar_linha(numero, {"tipo": "ERRO"}, {}))
                    except Exception as e:
                        msg = str(e).lower()
                        if any(k in msg for k in _ERROS_SITE) and tentativa < 3:
                            print(f"   ⏳ Erro de site — reconectando ({tentativa}/3)...")
                            time.sleep(10)
                            try:
                                projudi.login(page, cpf, senha, log)
                                url_2g, url_1g = projudi.get_urls_busca(page, log)
                            except Exception:
                                pass
                        else:
                            print(f"   ❌ Erro: {e}")
                            linhas.append(_montar_linha(numero, {"tipo": "ERRO"}, {}))
                            break

        except KeyboardInterrupt:
            print("\n\n(Interrompido pelo usuário — salvando resultados parciais...)")
        finally:
            try:
                browser.close()
            except Exception:
                pass

    # Saída
    _imprimir_resumo(linhas)

    nome_arquivo = f"VALIDACAO_PROMPT_{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx"
    caminho_xlsx = str(OUTPUT / nome_arquivo)
    _salvar_xlsx_validacao(linhas, caminho_xlsx)

    print(f"\n✅ Concluído. {len(linhas)} processo(s) processados.")
    print(f"   Abra o XLSX e preencha as colunas:")
    print(f"   • VALIDAÇÃO        → CORRETO ou ERRADO")
    print(f"   • ERRO — STATUS    → status correto (se errado)")
    print(f"   • ERRO — MATÉRIA   → matéria correta (se errada)")
    print(f"   • OBSERVAÇÃO       → descrição do erro\n")


if __name__ == "__main__":
    main()
